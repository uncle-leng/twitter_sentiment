from datetime import datetime
from django.shortcuts import render
from django.http import Http404, HttpResponseRedirect, HttpResponse
from couchdb.client import Server
from openpyxl import load_workbook
import json
from .util import parse, parse_travel
# Create your views here.



def index(request):
    return render(request, 'index.html')

def twitter_user(request):
    return render(request, 'twitter_user.html')

def words(request):
    return render(request,'words.html')

def cities(request):
    if 'sentiment' not in request.GET:
        return render(request, 'cities.html')
    if request.GET['sentiment']:
        res = []
        wb = load_workbook("twitter_backend/static/SentimentScoreByCities.xlsx")
        sheet = wb['Sheet1']
        for row in sheet.rows:
            tmp = []
            for cell in row:
                tmp.append(cell.value)
            res.append(tmp[1: -1])
        return HttpResponse(json.dumps({'sentiment_data': res[1:]}))



def words(request):
    if 'sentiment' not in request.GET:
        return render(request, 'words.html')
    else:
        res = []
        wb = load_workbook("twitter_backend/static/UserWordsofInterestV2.xlsx")
        sheet = wb['Sheet1']
        tmp = []
        for row in sheet.rows:
            tmp_list = []
            for cell in row:
                if isinstance(cell.value, datetime):
                    tmp_list.append(cell.value.strftime('%Y-%m-%d'))
                else:
                    tmp_list.append(cell.value)
            #tuple=[row[0],row[1]]
            tmp.append(tmp_list)
        return HttpResponse(json.dumps({'WordsWithWeight': tmp}))

def map_aurin(request):
    if 'sentiment' not in request.GET:
        return render(request, 'map_aurin.html')
    else:
        res = []
        wb = load_workbook("twitter_backend/static/AurinDataWithSentimentScore.xlsx")
        sheet = wb['Sheet1']
        for col in sheet.columns:
            tmp = []
            for cell in col:
                tmp.append(cell.value)
            res.append(tmp[1:])
        return HttpResponse(json.dumps({'sentiment_data': res[1:]}))


def map_jpg(request):
    imagepath = "twitter_backend/static/img/map.jpg"
    image_data = open(imagepath, "rb").read()
    return HttpResponse(image_data, content_type="image/jpg")

def user_detail(request):
    if 'detail' not in request.GET:
        return render(request, 's4.html')
    elif request.GET['detail'] == '1':
        with open("twitter_backend/static/S4_1_Time&Tweets.txt", 'r') as f:
            for line in f:
                dic = eval(line)
                dic_sort = sorted(dic.items(), key=lambda x: x[0][0])
                f.close()
                return HttpResponse(json.dumps(dic_sort))

    elif request.GET['detail'] == '2':
        res1 = parse("twitter_backend/static/S4_2_LocationHeatMap.txt")
        res2 = parse_travel("twitter_backend/static/S4_3_TravelMap.txt")
        return HttpResponse(json.dumps({"result1": res1, "result2": res2}))




def test(request):
    server = Server('http://43.240.96.132:5984/')
    documents = server['testdb']
    if request.method == 'GET':
        return render(request, 'test.html')
    if request.method == 'POST':
        x = {'haha': 'sdfgdfgf'}
        return HttpResponse(json.dumps(x))




